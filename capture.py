
import cv2
from datetime import datetime
import openpyxl
import os
 
# Predefined width and height (modify as needed)
width = 640
height = 640
 
# Check if the Excel file already exists
excel_filename = "frame_details.xlsx"
if os.path.exists(excel_filename):
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    # Create headers if this is a new file
    sheet.cell(row=1, column=1).value = "Filename"
    sheet.cell(row=1, column=2).value = "Resolution"
    sheet.cell(row=1, column=3).value = "Aspect Ratio"
    sheet.cell(row=1, column=4).value = "Created"
    sheet.cell(row=1, column=5).value = "cirA"
    sheet.cell(row=1, column=6).value = "cirB"
    sheet.cell(row=1, column=7).value = "cirC"
    sheet.cell(row=1, column=8).value = "CirD"
   
    for column_index, desired_width in enumerate([15, 15, 15, 23, 15], start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = desired_width
 
# Find the next available row in the sheet
next_row = sheet.max_row + 1
 
start_frame_count = int(input("Enter the starting frame number: "))
frame_count = start_frame_count
camera = cv2.VideoCapture(0)
while True:
    ret, frame = camera.read()
    if not ret:  # Check if frame is None
        print("Failed to capture frame. Exiting...")
        break
    cv2.imshow('frame', frame)
    # Check for key press
    key = cv2.waitKey(0)
    if key == ord('q'):  # Quit if 'q' is pressed
        break
    elif key == ord('c'):  # Capture image if 'c' is pressed
        frame_count += 1
        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"frame_{frame_count}.jpg"
        resolution = f"{frame.shape[1]}x{frame.shape[0]}"
        aspect_ratio = frame.shape[1] / frame.shape[0] if frame.shape[0] > 0 else "N/A"
        # Write details to Excel sheet
        sheet.cell(row=next_row, column=1).value = filename
        sheet.cell(row=next_row, column=2).value = resolution
        sheet.cell(row=next_row, column=3).value = aspect_ratio
        sheet.cell(row=next_row, column=4).value = current_datetime
        # Placeholder values for cir1, cir2, cir3, cir4, and description
        # sheet.cell(row=next_row, column=5).value = cir1
        # sheet.cell(row=next_row, column=6).value = cir2
        # sheet.cell(row=next_row, column=7).value = cir3
        # sheet.cell(row=next_row, column=8).value = cir4
       
        # Save the frame
        cv2.imwrite(filename, frame)
 
        print(f"Frame captured: {filename}")
        wb.save(excel_filename)
 
        next_row += 1
 
camera.release()
cv2.destroyAllWindows()
 